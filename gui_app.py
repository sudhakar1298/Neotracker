import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import time
import os
import imaplib
import email
import sqlite3
import logging
from datetime import datetime
from email.header import decode_header
from dotenv import load_dotenv, set_key
from openpyxl import load_workbook
from plyer import notification
import ollama

# --- CONFIGURATION ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")
ENV_FILE = ".env"

# Ensure .env exists
if not os.path.exists(ENV_FILE):
    with open(ENV_FILE, "w") as f:
        f.write("EMAIL_USER=\nEMAIL_PASS=\nIMAP_SERVER=imap.gmail.com\nTARGET_ID=\nCHECK_INTERVAL=30\n")

load_dotenv(ENV_FILE)

# --- BACKEND LOGIC ---
class Database:
    def __init__(self, db_name="history.db"):
        self.conn = sqlite3.connect(db_name, check_same_thread=False)
        self.create_table()

    def create_table(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                company TEXT,
                source TEXT,
                details TEXT
            )
        """)
        self.conn.commit()

    def log_match(self, company, source, details):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute("INSERT INTO matches (timestamp, company, source, details) VALUES (?, ?, ?, ?)",
                          (ts, company, source, details))
        self.conn.commit()

    def get_all(self):
        cursor = self.conn.execute("SELECT timestamp, company, source, details FROM matches ORDER BY id DESC")
        return cursor.fetchall()

class MailWorker:
    def __init__(self, log_callback, success_callback):
        self.running = False
        self.log = log_callback
        self.on_success = success_callback
        self.db = Database()
        
    def get_config(self, key):
        return os.getenv(key, "")

    def extract_company(self, subject, body):
        try:
            prompt = f"Extract company name from subject: '{subject}' and body snippet: '{body[:500]}'. Reply ONLY with name."
            response = ollama.chat(model="llama3", messages=[{"role": "user", "content": prompt}])
            return response["message"]["content"].strip()
        except:
            return "Unknown Company"

    def check_excel(self, path, target_id):
        try:
            wb = load_workbook(path, data_only=True)
            text = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_str = " ".join([str(c) for c in row if c])
                    if target_id in row_str:
                        return True, "Exact Match"
                    text += row_str + "\n"
            
            # LLM Check if no exact match
            prompt = f"My ID is {target_id}. Check this list: {text[:2000]}. Does it contain my ID? YES/NO only."
            res = ollama.chat(model="llama3", messages=[{"role": "user", "content": prompt}])
            return "YES" in res["message"]["content"].upper(), "LLM Match"
        except Exception as e:
            self.log(f"Excel Error: {e}")
            return False, ""

    def run_check(self):
        email_user = self.get_config("EMAIL_USER")
        email_pass = self.get_config("EMAIL_PASS")
        target_id = self.get_config("TARGET_ID")
        
        if not (email_user and email_pass and target_id):
            self.log("ERROR: Missing Credentials in Settings!")
            self.running = False
            return

        self.log("--- Starting Mail Check Cycle ---")
        
        try:
            mail = imaplib.IMAP4_SSL(self.get_config("IMAP_SERVER"))
            mail.login(email_user, email_pass)
            mail.select("inbox")
            
            today = datetime.now().strftime("%d-%b-%Y")
            # Using generic search for demonstration, refine sender list as needed
            status, data = mail.search(None, f'(UNSEEN SINCE "{today}")')

            if not data or not data[0]:
                self.log("No new unread mails.")
            else:
                for num in data[0].split():
                    _, msg_data = mail.fetch(num, '(RFC822)')
                    msg = email.message_from_bytes(msg_data[0][1])
                    
                    subject = decode_header(msg["Subject"])[0][0]
                    if isinstance(subject, bytes): subject = subject.decode()
                    
                    self.log(f"Checking: {subject[:30]}...")

                    # Check Body
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body += part.get_payload(decode=True).decode(errors="ignore")
                    else:
                        body = msg.get_payload(decode=True).decode(errors="ignore")

                    company = self.extract_company(subject, body)

                    if target_id in (subject + body):
                        self.log(f"MATCH FOUND in Body: {company}")
                        self.db.log_match(company, "Email Body", subject)
                        self.on_success(company)
                    
                    # Check Attachments
                    for part in msg.walk():
                        if part.get_content_disposition() == "attachment":
                            fname = part.get_filename()
                            if fname and fname.endswith(".xlsx"):
                                path = f"attachments/{fname}"
                                os.makedirs("attachments", exist_ok=True)
                                with open(path, "wb") as f: f.write(part.get_payload(decode=True))
                                
                                match, reason = self.check_excel(path, target_id)
                                if match:
                                    self.log(f"MATCH FOUND in Excel: {company}")
                                    self.db.log_match(company, "Excel File", f"{fname} ({reason})")
                                    self.on_success(company)
                    
                    mail.store(num, '+FLAGS', '\\Seen')

            mail.logout()
        except Exception as e:
            self.log(f"Connection Error: {e}")

        self.log("--- Cycle Finished ---")

# --- GUI FRONTEND ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Placement ID Watcher")
        self.geometry("900x600")
        
        # Layout
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        self.logo_label = ctk.CTkLabel(self.sidebar, text="ID WATCHER", font=ctk.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.btn_dashboard = ctk.CTkButton(self.sidebar, text="Dashboard", command=self.show_dashboard)
        self.btn_dashboard.grid(row=1, column=0, padx=20, pady=10)
        
        self.btn_history = ctk.CTkButton(self.sidebar, text="History", command=self.show_history)
        self.btn_history.grid(row=2, column=0, padx=20, pady=10)

        self.btn_settings = ctk.CTkButton(self.sidebar, text="Settings", command=self.show_settings)
        self.btn_settings.grid(row=3, column=0, padx=20, pady=10)

        # Main Area
        self.main_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, sticky="nsew")

        self.create_dashboard()
        self.create_history()
        self.create_settings()
        
        self.worker_thread = None
        self.worker_running = False
        
        self.show_dashboard()

    def create_dashboard(self):
        self.dash_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        
        self.status_label = ctk.CTkLabel(self.dash_frame, text="Status: Stopped", font=("Arial", 16))
        self.status_label.pack(pady=10)

        self.btn_toggle = ctk.CTkButton(self.dash_frame, text="START MONITORING", 
                                        fg_color="green", hover_color="darkgreen", height=50,
                                        command=self.toggle_monitoring)
        self.btn_toggle.pack(pady=10, fill="x", padx=50)

        self.log_box = ctk.CTkTextbox(self.dash_frame, height=400)
        self.log_box.pack(pady=10, padx=20, fill="both", expand=True)
        self.log_box.configure(state="disabled")

    def create_history(self):
        self.hist_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        
        cols = ("Timestamp", "Company", "Source", "Details")
        self.tree = ttk.Treeview(self.hist_frame, columns=cols, show="headings", height=20)
        
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)
            
        self.tree.pack(fill="both", expand=True, padx=20, pady=20)
        
        btn_refresh = ctk.CTkButton(self.hist_frame, text="Refresh", command=self.load_history_data)
        btn_refresh.pack(pady=10)

    def create_settings(self):
        self.set_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        
        fields = ["EMAIL_USER", "EMAIL_PASS", "TARGET_ID", "CHECK_INTERVAL"]
        self.entries = {}
        
        for i, field in enumerate(fields):
            lbl = ctk.CTkLabel(self.set_frame, text=field.replace("_", " ").title() + ":")
            lbl.grid(row=i, column=0, padx=20, pady=10, sticky="w")
            
            entry = ctk.CTkEntry(self.set_frame, width=300)
            if "PASS" in field: entry.configure(show="*")
            entry.grid(row=i, column=1, padx=20, pady=10)
            entry.insert(0, os.getenv(field, ""))
            self.entries[field] = entry

        btn_save = ctk.CTkButton(self.set_frame, text="Save Settings", command=self.save_settings)
        btn_save.grid(row=len(fields), column=1, pady=20)

    # --- ACTIONS ---
    def show_dashboard(self): self.switch_frame(self.dash_frame)
    def show_history(self): 
        self.load_history_data()
        self.switch_frame(self.hist_frame)
    def show_settings(self): self.switch_frame(self.set_frame)
    
    def switch_frame(self, frame):
        for w in self.main_frame.winfo_children(): w.pack_forget()
        frame.pack(fill="both", expand=True)

    def log(self, message):
        def _update():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _update)

    def on_match_found(self, company):
        notification.notify(title="Shortlist Found!", message=f"Company: {company}", timeout=10)
        self.log(f"!!! ALERT: FOUND MATCH FOR {company} !!!")

    def toggle_monitoring(self):
        if not self.worker_running:
            self.worker_running = True
            self.btn_toggle.configure(text="STOP MONITORING", fg_color="red", hover_color="darkred")
            self.status_label.configure(text="Status: Running", text_color="green")
            
            threading.Thread(target=self.bg_loop, daemon=True).start()
        else:
            self.worker_running = False
            self.btn_toggle.configure(text="START MONITORING", fg_color="green", hover_color="darkgreen")
            self.status_label.configure(text="Status: Stopped", text_color="white")
            self.log("Stopping after current cycle...")

    def bg_loop(self):
        worker = MailWorker(self.log, self.on_match_found)
        while self.worker_running:
            worker.run_check()
            # Sleep in small chunks to allow faster stopping
            for _ in range(int(os.getenv("CHECK_INTERVAL", 30))):
                if not self.worker_running: break
                time.sleep(1)

    def save_settings(self):
        for key, entry in self.entries.items():
            set_key(ENV_FILE, key, entry.get())
        load_dotenv(ENV_FILE, override=True)
        messagebox.showinfo("Saved", "Settings saved successfully! Restart monitoring to apply.")

    def load_history_data(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        db = Database()
        for row in db.get_all():
            self.tree.insert("", "end", values=row)

if __name__ == "__main__":
    app = App()
    app.mainloop()  