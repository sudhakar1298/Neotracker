import customtkinter as ctk
from tkinter import ttk, messagebox
import threading
import time
import os
import imaplib
import email
import sqlite3
import re
import sys
import winreg
import requests
import json
from datetime import datetime
from email.header import decode_header
from dotenv import load_dotenv, set_key
from openpyxl import load_workbook
from winotify import Notification, audio
import pystray
from PIL import Image, ImageDraw

# --- 🎨 VISUAL DESIGN GUIDELINES (CYBERPUNK THEME) ---
THEME = {
    "bg_primary": "#0A0A0A",      # Deep space black
    "bg_secondary": "#1E1E1E",    # Dark gray
    "bg_card": "#161B22",         # Slightly lighter for inputs
    "accent_blue": "#00D4FF",     # Electric blue
    "accent_green": "#00FF88",    # Cyber green
    "accent_red": "#FF4444",      # Neon red
    "text_primary": "#FFFFFF",    # Pure white
    "text_secondary": "#B0B0B0",  # Light gray
    "border_color": "#30363D",    # Subtle border
    "font_code": ("Consolas", 13),
    "font_ui": ("Arial", 13),
    "font_header": ("Arial", 20, "bold")
}

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# ... imports ...

# --- SAFER CONFIGURATION LOADING ---
# Save the .env file in the User's Home folder (e.g., C:\Users\Name\.placement_watcher.env)
# This prevents "Permission Denied" errors when the app runs on startup.
USER_HOME = os.path.expanduser("~")
ENV_FILE = os.path.join(USER_HOME, ".placement_watcher.env")

try:
    if not os.path.exists(ENV_FILE):
        with open(ENV_FILE, "w") as f:
            f.write("EMAIL_USER=\nEMAIL_PASS=\nIMAP_SERVER=imap.gmail.com\nTARGET_ID=\nCHECK_INTERVAL=30\nAI_MODEL=llama3\n")
    load_dotenv(ENV_FILE)
except PermissionError:
    print("⚠️ Config Permission Error: The file is locked or read-only.")
except Exception as e:
    print(f"⚠️ Config Error: {e}")

# ... Rest of the code ...# --- BACKEND LOGIC ---
class Database:
    def __init__(self, db_name="history.db"):
        self.conn = sqlite3.connect(db_name, check_same_thread=False)
        self.create_table()

    def create_table(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                company TEXT,
                source TEXT,
                details TEXT
            )
        """)
        self.conn.commit()

    def log_match(self, company, source, details):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute("INSERT INTO matches (timestamp, company, source, details) VALUES (?, ?, ?, ?)",
                          (ts, company, source, details))
        self.conn.commit()

    def get_all(self):
        cursor = self.conn.execute("SELECT timestamp, company, source, details FROM matches ORDER BY id DESC")
        return cursor.fetchall()

    def clear_all(self):
        self.conn.execute("DELETE FROM matches")
        self.conn.commit()

class MailWorker:
    def __init__(self, log_callback, success_callback, update_ai_status):
        self.running = False
        self.log = log_callback
        self.on_success = success_callback
        self.update_ai_status = update_ai_status
        self.db = Database()
        self.ai_available = False
        
    def get_config(self, key): return os.getenv(key, "")

    def check_ollama_status(self):
        try:
            response = requests.get("http://localhost:11434/", timeout=1)
            if response.status_code == 200:
                self.ai_available = True
                self.update_ai_status(True)
                return True
        except:
            pass
        self.ai_available = False
        self.update_ai_status(False)
        return False

    def ask_ollama(self, prompt):
        model = self.get_config("AI_MODEL") or "llama3"
        url = "http://localhost:11434/api/generate"
        data = {"model": model, "prompt": prompt, "stream": False}
        try:
            resp = requests.post(url, json=data, timeout=10)
            return resp.json().get("response", "").strip()
        except Exception as e:
            self.log(f"⚠️ AI Failed: {e}")
            return None

    def extract_company(self, subject):
        if self.ai_available:
            prompt = f"Extract ONLY the company name from this email subject: '{subject}'. Do not output anything else. If no company is found, return 'Unknown'."
            result = self.ask_ollama(prompt)
            if result:
                return result.replace('"', '').replace("'", "")

        # Regex Fallback
        clean = subject.replace("Fwd:", "").replace("Re:", "").strip()
        junk = [r"Shortlist(ed)?", r"Selected", r"regarding", r"Placement", r"Hiring", r"Online Test", r"Interview", r"Round", r"Batch", r"202\d"]
        for j in junk:
            clean = re.sub(j, "", clean, flags=re.IGNORECASE)
        return clean.strip(" -:|")[:30] if clean.strip() else "Unknown"

    def check_excel_simple(self, path, target_id):
        try:
            wb = load_workbook(path, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_str = " ".join([str(c) for c in row if c]).upper()
                    if target_id.upper() in row_str:
                        return True, "Exact Match in Excel"
            return False, ""
        except Exception as e:
            self.log(f"Excel Error: {e}")
            return False, ""

    def run_check(self):
        email_user = self.get_config("EMAIL_USER")
        email_pass = self.get_config("EMAIL_PASS")
        target_id = self.get_config("TARGET_ID")
        
        if not (email_user and email_pass and target_id):
            self.log("❌ ERROR: Credentials missing.")
            self.running = False
            return

        self.check_ollama_status()
        self.log(f">>> SCANNING... (AI Mode: {'ON' if self.ai_available else 'OFF'})")

        try:
            mail = imaplib.IMAP4_SSL(self.get_config("IMAP_SERVER"))
            mail.login(email_user, email_pass)
            mail.select("inbox")
            
            today = datetime.now().strftime("%d-%b-%Y")
            status, data = mail.search(None, f'(UNSEEN SINCE "{today}")')

            if not data or not data[0]:
                self.log("No new emails.")
            else:
                for num in data[0].split():
                    _, msg_data = mail.fetch(num, '(RFC822)')
                    msg = email.message_from_bytes(msg_data[0][1])
                    
                    subject_bytes = msg["Subject"]
                    try:
                        decoded = decode_header(subject_bytes)[0]
                        subject = decoded[0]
                        if isinstance(subject, bytes): subject = subject.decode(decoded[1] or 'utf-8')
                    except: subject = str(subject_bytes)

                    company = self.extract_company(str(subject))
                    self.log(f"Checking: {company}...")

                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body += part.get_payload(decode=True).decode(errors="ignore")
                    else:
                        body = msg.get_payload(decode=True).decode(errors="ignore")

                    if target_id in (str(subject) + body):
                        self.on_success(company)
                        self.db.log_match(company, "Email Body", subject)
                    
                    for part in msg.walk():
                        if part.get_content_disposition() == "attachment":
                            fname = part.get_filename()
                            if fname and fname.endswith(".xlsx"):
                                os.makedirs("attachments", exist_ok=True)
                                path = f"attachments/{fname}"
                                with open(path, "wb") as f: f.write(part.get_payload(decode=True))
                                
                                match, reason = self.check_excel_simple(path, target_id)
                                if match:
                                    self.on_success(company)
                                    self.db.log_match(company, "Excel", f"{fname} ({reason})")
                                    
                    mail.store(num, '+FLAGS', '\\Seen')
            mail.logout()
        except Exception as e:
            self.log(f"Connection Error: {e}")

# --- GUI ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ID WATCHER (Hybrid AI)")
        self.geometry("900x600")
        self.configure(fg_color=THEME["bg_primary"])
        
        # INTERCEPT CLOSE BUTTON
        self.protocol("WM_DELETE_WINDOW", self.minimize_to_tray)

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color=THEME["bg_secondary"])
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        # Logo Area
        ctk.CTkLabel(self.sidebar, text="ID WATCHER", font=THEME["font_header"], text_color=THEME["accent_blue"]).grid(row=0, column=0, pady=(40,20), padx=20)
        ctk.CTkLabel(self.sidebar, text="SYSTEM v2.0", font=("Consolas", 10), text_color="grey").grid(row=1, column=0, pady=(0,30))
        
        self.nav_btns = {}
        for i, (name, cmd) in enumerate([("Dashboard", self.show_dashboard), ("History", self.show_history), ("Settings", self.show_settings)]):
            btn = ctk.CTkButton(self.sidebar, text=f"  {name}", command=cmd, 
                              fg_color="transparent", hover_color=THEME["bg_card"], 
                              anchor="w", font=THEME["font_ui"], height=40, corner_radius=8)
            btn.grid(row=i+2, column=0, padx=15, pady=5, sticky="ew")
            self.nav_btns[name] = btn

        # Main Area
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, sticky="nsew", padx=30, pady=30)
        
        self.worker_running = False
        self.create_frames()
        self.show_dashboard()

        # Startup
        self.after(100, self.check_auto_start)

        # Setup System Tray
        threading.Thread(target=self.setup_tray, daemon=True).start()

    # --- SYSTEM TRAY LOGIC ---
    def setup_tray(self):
        image = Image.new('RGB', (64, 64), color=(0, 212, 255))
        draw = ImageDraw.Draw(image)
        draw.rectangle([16, 16, 48, 48], fill="black")
        
        menu = (pystray.MenuItem('Show', self.show_window_from_tray), 
                pystray.MenuItem('Quit', self.quit_app))
        self.tray_icon = pystray.Icon("name", image, "Placement Watcher", menu)
        self.tray_icon.run()

    def minimize_to_tray(self):
        self.withdraw()  # Hide the window
        # Notification to tell user it's still running
        toast = Notification(app_id="Placement Watcher", title="Minimized to Tray", msg="I am still scanning in the background.", duration="short")
        toast.show()

    def show_window_from_tray(self, icon, item):
        self.after(0, self.deiconify) # Show the window

    def quit_app(self, icon, item):
        self.worker_running = False
        self.tray_icon.stop()
        self.quit()

    # --- APP LOGIC ---
    def check_auto_start(self):
        if os.getenv("EMAIL_USER") and os.getenv("EMAIL_PASS"):
            self.log("Auto-starting background monitor...")
            self.toggle_monitoring()
        else:
            self.log("Please configure Settings to start.")

    def create_frames(self):
        # Dashboard
        self.dash_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        
        # Status Bar
        status_bar = ctk.CTkFrame(self.dash_frame, fg_color=THEME["bg_card"], corner_radius=12, border_width=1, border_color=THEME["border_color"])
        status_bar.pack(fill="x", pady=(0, 20))
        
        self.ai_status_label = ctk.CTkLabel(status_bar, text="● AI: CHECKING", text_color="grey", font=("Consolas", 12, "bold"))
        self.ai_status_label.pack(side="left", padx=20, pady=15)

        self.status_btn = ctk.CTkButton(status_bar, text="START MONITORING", fg_color=THEME["accent_blue"], hover_color="#00B8E6",
                                      text_color="black", height=32, font=("Arial", 12, "bold"), command=self.toggle_monitoring)
        self.status_btn.pack(side="right", padx=20, pady=15)
        
        # Log Box
        ctk.CTkLabel(self.dash_frame, text="LIVE LOGS", font=("Consolas", 12, "bold"), text_color="grey").pack(anchor="w", pady=(0, 5))
        self.log_box = ctk.CTkTextbox(self.dash_frame, fg_color=THEME["bg_secondary"], 
                                    text_color=THEME["accent_green"], font=THEME["font_code"],
                                    corner_radius=12, border_width=1, border_color=THEME["border_color"])
        self.log_box.pack(fill="both", expand=True)

        # History Frame
        self.hist_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        
        header = ctk.CTkFrame(self.hist_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(header, text="MATCH HISTORY", font=THEME["font_header"], text_color="white").pack(side="left")
        
        ctk.CTkButton(header, text="CLEAR ALL", command=self.clear_history, 
                    fg_color=THEME["bg_card"], hover_color=THEME["accent_red"], 
                    text_color="white", width=100).pack(side="right")

        table_header = ctk.CTkFrame(self.hist_frame, fg_color=THEME["bg_card"], height=35, corner_radius=8)
        table_header.pack(fill="x", pady=(0,5))
        
        table_header.grid_columnconfigure(0, weight=2)
        table_header.grid_columnconfigure(1, weight=3)
        table_header.grid_columnconfigure(2, weight=2)
        
        ctk.CTkLabel(table_header, text="TIME", font=("Consolas", 11, "bold"), text_color="grey").grid(row=0, column=0, padx=10, pady=8, sticky="w")
        ctk.CTkLabel(table_header, text="COMPANY", font=("Consolas", 11, "bold"), text_color="grey").grid(row=0, column=1, padx=10, pady=8, sticky="w")
        ctk.CTkLabel(table_header, text="SOURCE", font=("Consolas", 11, "bold"), text_color="grey").grid(row=0, column=2, padx=10, pady=8, sticky="w")

        self.tree_scroll = ctk.CTkScrollableFrame(self.hist_frame, fg_color="transparent")
        self.tree_scroll.pack(fill="both", expand=True)

        # Settings
        self.set_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        ctk.CTkLabel(self.set_frame, text="CONFIGURATION", font=THEME["font_header"], text_color="white").grid(row=0, column=0, sticky="w", pady=(0,30))
        self.entries = {}
        fields = ["EMAIL_USER", "EMAIL_PASS", "TARGET_ID", "CHECK_INTERVAL", "AI_MODEL"]
        for i, f in enumerate(fields):
            ctk.CTkLabel(self.set_frame, text=f.replace("_", " "), font=("Arial", 12, "bold"), text_color="grey").grid(row=i*2+1, column=0, sticky="w", pady=(10,5))
            ent = ctk.CTkEntry(self.set_frame, width=400, height=40, fg_color=THEME["bg_card"], border_color=THEME["border_color"], text_color="white")
            if "PASS" in f: ent.configure(show="•")
            ent.insert(0, os.getenv(f, ""))
            ent.grid(row=i*2+2, column=0, sticky="w")
            self.entries[f] = ent
            
        self.startup_var = ctk.BooleanVar(value=self.check_startup_registry())
        ctk.CTkCheckBox(self.set_frame, text="Run on Windows Startup", variable=self.startup_var, command=self.toggle_startup_registry, fg_color=THEME["accent_blue"], hover_color=THEME["accent_blue"]).grid(row=12, column=0, pady=20, sticky="w")
        ctk.CTkButton(self.set_frame, text="SAVE SETTINGS", command=self.save_settings, fg_color=THEME["accent_green"], text_color="black", font=("Arial", 13, "bold"), height=45, width=400).grid(row=13, column=0)
        
        note_text = "💡 TIP: If 'Ollama' is running in the background, the app will automatically\ndetect it and switch to AI Mode for smarter company detection."
        ctk.CTkLabel(self.set_frame, text=note_text, font=("Consolas", 11), text_color="grey", justify="left").grid(row=14, column=0, pady=(20,0), sticky="w")

        ctk.CTkLabel(self.set_frame, text="⚡ System built by Sudhakar", font=("Consolas", 12, "bold"), text_color=THEME["accent_blue"]).grid(row=15, column=0, pady=(30,0))

    # --- ACTIONS ---
    def update_ai_indicator(self, active):
        color = THEME["accent_green"] if active else "grey"
        text = "● AI: ONLINE" if active else "● AI: OFFLINE (Using Regex)"
        self.ai_status_label.configure(text=text, text_color=color)

    def switch_frame(self, frame):
        for f in [self.dash_frame, self.hist_frame, self.set_frame]: f.pack_forget()
        frame.pack(fill="both", expand=True)

    def show_dashboard(self): self.switch_frame(self.dash_frame)
    def show_history(self): 
        self.load_history()
        self.switch_frame(self.hist_frame)
    def show_settings(self): self.switch_frame(self.set_frame)

    def log(self, msg):
        self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M')}] {msg}\n")
        self.log_box.see("end")

    def toggle_monitoring(self):
        if not self.worker_running:
            self.worker_running = True
            self.status_btn.configure(text="STOP MONITORING", fg_color=THEME["accent_red"], hover_color="#CC0000")
            threading.Thread(target=self.bg_loop, daemon=True).start()
        else:
            self.worker_running = False
            self.status_btn.configure(text="START MONITORING", fg_color=THEME["accent_blue"], hover_color="#00B8E6")
            self.log("Stopping...")

    def bg_loop(self):
        def send_alert(company_name):
            toast = Notification(app_id="Placement Watcher", title="MATCH FOUND!", msg=f"Company: {company_name}", duration="long")
            toast.show()

        worker = MailWorker(self.log, send_alert, self.update_ai_indicator)
        while self.worker_running:
            worker.run_check()
            for _ in range(int(os.getenv("CHECK_INTERVAL", 30))):
                if not self.worker_running: break
                time.sleep(1)

    def load_history(self):
        for w in self.tree_scroll.winfo_children(): w.destroy()
        rows = Database().get_all()
        
        if not rows:
            ctk.CTkLabel(self.tree_scroll, text="No history found.", text_color="grey").pack(pady=20)
            return

        for row in rows:
            row_frame = ctk.CTkFrame(self.tree_scroll, fg_color=THEME["bg_card"], corner_radius=6)
            row_frame.pack(fill="x", pady=2)
            row_frame.grid_columnconfigure(0, weight=2)
            row_frame.grid_columnconfigure(1, weight=3)
            row_frame.grid_columnconfigure(2, weight=2)

            ctk.CTkLabel(row_frame, text=row[0], font=("Consolas", 11), text_color="grey").grid(row=0, column=0, padx=10, pady=10, sticky="w")
            ctk.CTkLabel(row_frame, text=row[1], font=("Arial", 12, "bold"), text_color=THEME["accent_blue"]).grid(row=0, column=1, padx=10, pady=10, sticky="w")
            ctk.CTkLabel(row_frame, text=row[2], font=("Arial", 11), text_color="white").grid(row=0, column=2, padx=10, pady=10, sticky="w")

    def clear_history(self):
        if messagebox.askyesno("Confirm", "Are you sure you want to delete all history?"):
            Database().clear_all()
            self.load_history()

    def save_settings(self):
        for k, v in self.entries.items(): set_key(ENV_FILE, k, v.get())
        load_dotenv(ENV_FILE, override=True)
        messagebox.showinfo("Saved", "Settings Updated.")

    def check_startup_registry(self):
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_READ)
            winreg.QueryValueEx(key, "PlacementWatcher")
            winreg.CloseKey(key)
            return True
        except: return False

    def toggle_startup_registry(self):
        app_path = os.path.abspath(sys.argv[0])
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_SET_VALUE)
        if self.startup_var.get():
            winreg.SetValueEx(key, "PlacementWatcher", 0, winreg.REG_SZ, f'"{app_path}"')
        else:
            try: winreg.DeleteValue(key, "PlacementWatcher")
            except: pass
        winreg.CloseKey(key)

if __name__ == "__main__":
    app = App()
    app.mainloop()