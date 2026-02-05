import customtkinter as ctk
from tkinter import ttk, messagebox
import threading
import time
import os
import imaplib
import email
import sqlite3
from datetime import datetime
from email.header import decode_header
from dotenv import load_dotenv, set_key
from openpyxl import load_workbook
from plyer import notification
import ollama

# --- THEME & MAC OPTIMIZATION ---
THEME = {
    "bg_primary": "#0A0A0A",
    "bg_secondary": "#1E1E1E",
    "accent_blue": "#00D4FF",
    "accent_green": "#00FF88",
    "accent_purple": "#8B5FBF",
    "text_primary": "#FFFFFF",
    "text_secondary": "#B0B0B0",
    "success": "#00FF41",
    "warning": "#FFB800",
    "error": "#FF4444",
    "font_code": ("JetBrains Mono", "Menlo", "Monaco", "Courier"),
    "font_ui": (".AppleSystemUIFont", "Inter", "Helvetica", "Arial")
}

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

ENV_FILE = ".env"
if not os.path.exists(ENV_FILE):
    with open(ENV_FILE, "w") as f:
        f.write("EMAIL_USER=\nEMAIL_PASS=\nIMAP_SERVER=imap.gmail.com\nTARGET_ID=\nCHECK_INTERVAL=30\n")
load_dotenv(ENV_FILE)

# --- BACKEND: DATABASE ---
class Database:
    def __init__(self, db_name="history.db"):
        self.db_path = os.path.join(os.path.dirname(__file__), db_name)
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.create_table()

    def create_table(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                company TEXT,
                source TEXT,
                details TEXT
            )
        """)
        self.conn.commit()

    def log_match(self, company, source, details):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute("INSERT INTO matches (timestamp, company, source, details) VALUES (?, ?, ?, ?)",
                          (ts, company, source, details))
        self.conn.commit()

    def get_all(self):
        try:
            cursor = self.conn.execute("SELECT timestamp, company, source, details FROM matches ORDER BY id DESC")
            return cursor.fetchall()
        except: return []

# --- BACKEND: MAIL WORKER WITH OLLAMA FALLBACK ---
class MailWorker:
    def __init__(self, log_callback, success_callback):
        self.log = log_callback
        self.on_success = success_callback
        self.db = Database()
        self.ai_enabled = self._verify_ollama()
        
    def _verify_ollama(self):
        """Checks if Ollama service is running and Llama3 is present."""
        try:
            models_info = ollama.list()
            # Support both 'llama3' and 'llama3:latest' naming
            available = any("llama3" in m['name'] for m in models_info.get('models', []))
            if available:
                self.log("SYSTEM: AI Engine Active (Llama3).")
                return True
            else:
                self.log("SYSTEM: Ollama found, but 'llama3' model missing. Using Keyword Mode.")
                return False
        except Exception:
            self.log("SYSTEM: Ollama not detected. Running in Keyword Mode (Safe).")
            return False

    def extract_company(self, subject, body):
        if self.ai_enabled:
            try:
                prompt = f"Extract company name from: '{subject}'. Reply ONLY with the name."
                res = ollama.chat(model="llama3", messages=[{"role": "user", "content": prompt}])
                return res["message"]["content"].strip()
            except: pass
        
        # Fallback keyword logic
        return subject.split(":")[0] if ":" in subject else "Detected Entity"

    def check_excel(self, path, target_id):
        try:
            wb = load_workbook(path, data_only=True)
            content_dump = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_str = " ".join([str(c) for c in row if c])
                    if target_id.lower() in row_str.lower():
                        return True, "Keyword Match"
                    content_dump += row_str + "\n"
            
            if self.ai_enabled:
                prompt = f"Does this list contain ID {target_id}? Context: {content_dump[:1000]}. Answer YES or NO."
                res = ollama.chat(model="llama3", messages=[{"role": "user", "content": prompt}])
                if "YES" in res["message"]["content"].upper():
                    return True, "AI Inference"
            
            return False, ""
        except Exception as e:
            self.log(f"Excel Error: {e}")
            return False, ""

    def run_check(self):
        user, pw = os.getenv("EMAIL_USER"), os.getenv("EMAIL_PASS")
        target = os.getenv("TARGET_ID")
        
        if not all([user, pw, target]):
            self.log("CRITICAL: Check credentials in Settings!")
            return

        try:
            mail = imaplib.IMAP4_SSL(os.getenv("IMAP_SERVER", "imap.gmail.com"))
            mail.login(user, pw)
            mail.select("inbox")
            
            # Scan only unseen mail from today
            date_tag = datetime.now().strftime("%d-%b-%Y")
            _, data = mail.search(None, f'(UNSEEN SINCE "{date_tag}")')

            if not data[0]:
                self.log("Scan: Clear. No new signals.")
            else:
                for num in data[0].split():
                    _, msg_data = mail.fetch(num, '(RFC822)')
                    msg = email.message_from_bytes(msg_data[0][1])
                    
                    # Decode Subject safely
                    raw_subj = decode_header(msg["Subject"])[0]
                    subj = raw_subj[0].decode(raw_subj[1] or 'utf-8') if isinstance(raw_subj[0], bytes) else raw_subj[0]
                    
                    self.log(f"Processing: {subj[:25]}...")
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body += part.get_payload(decode=True).decode(errors='ignore')
                    else:
                        body = msg.get_payload(decode=True).decode(errors='ignore')

                    company = self.extract_company(subj, body)

                    # Body Check
                    if target.lower() in (subj + body).lower():
                        self.log(f"MATCH: {company}")
                        self.db.log_match(company, "Email", subj)
                        self.on_success(company)
                    
                    # Attachment Check
                    for part in msg.walk():
                        if part.get_content_disposition() == "attachment":
                            fname = part.get_filename()
                            if fname and fname.endswith(".xlsx"):
                                os.makedirs("attachments", exist_ok=True)
                                path = f"attachments/{fname}"
                                with open(path, "wb") as f: f.write(part.get_payload(decode=True))
                                match, reason = self.check_excel(path, target)
                                if match:
                                    self.log(f"ATTACHMENT MATCH: {company} ({reason})")
                                    self.db.log_match(company, f"Excel: {fname}", reason)
                                    self.on_success(company)
                    
                    mail.store(num, '+FLAGS', '\\Seen')
            mail.logout()
        except Exception as e:
            self.log(f"Network: {e}")

# --- UI COMPONENTS ---
class CyberButton(ctk.CTkButton):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(corner_radius=8, font=(THEME["font_ui"][0], 13, "bold"), border_width=1,
                       border_color=THEME["accent_blue"], fg_color="transparent", text_color=THEME["accent_blue"])

class CyberEntry(ctk.CTkEntry):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(corner_radius=8, fg_color=THEME["bg_secondary"], border_color="#333",
                       text_color="white", font=(THEME["font_code"][0], 12))

# --- MAIN APPLICATION ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ID WATCHER // V1.0-MAC")
        self.geometry("1000x650")
        self.configure(fg_color=THEME["bg_primary"])
        
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color=THEME["bg_secondary"])
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        ctk.CTkLabel(self.sidebar, text="ID WATCHER", font=(THEME["font_code"][0], 22, "bold"), 
                     text_color=THEME["accent_blue"]).grid(row=0, column=0, padx=20, pady=40)

        self.create_nav_btn("DASHBOARD", self.show_dashboard, 1)
        self.create_nav_btn("HISTORY", self.show_history, 2)
        self.create_nav_btn("SETTINGS", self.show_settings, 3)

        self.status_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.status_frame.grid(row=10, column=0, pady=50)
        self.status_dot = ctk.CTkLabel(self.status_frame, text="●", font=("Arial", 24), text_color="grey")
        self.status_dot.pack(side="left", padx=5)
        self.status_text = ctk.CTkLabel(self.status_frame, text="OFFLINE", font=(THEME["font_code"][0], 12), text_color="grey")
        self.status_text.pack(side="left")

        # Main Container
        self.main_container = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.main_container.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)

        self.worker_running = False
        self.dash_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.hist_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.set_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        
        self.setup_dashboard()
        self.setup_history()
        self.setup_settings()
        self.show_dashboard()

    def create_nav_btn(self, text, command, row):
        btn = ctk.CTkButton(self.sidebar, text=text, command=command, font=(THEME["font_ui"][0], 14),
                            fg_color="transparent", text_color=THEME["text_secondary"], anchor="w", height=40, width=180)
        btn.grid(row=row, column=0, padx=20, pady=5)

    def setup_dashboard(self):
        ctk.CTkLabel(self.dash_frame, text="LIVE MONITOR", font=(THEME["font_code"][0], 20, "bold")).pack(anchor="w", pady=(0, 20))
        
        card = ctk.CTkFrame(self.dash_frame, fg_color=THEME["bg_secondary"], corner_radius=12)
        card.pack(fill="x", pady=10)

        self.btn_toggle = ctk.CTkButton(card, text="START MONITOR", font=(THEME["font_ui"][0], 15, "bold"),
                                        fg_color=THEME["accent_blue"], text_color="black", height=50, command=self.toggle_monitoring)
        self.btn_toggle.pack(padx=20, pady=20, fill="x")

        self.log_box = ctk.CTkTextbox(self.dash_frame, fg_color="#050505", text_color=THEME["success"], 
                                      font=(THEME["font_code"][0], 12), border_width=1, border_color="#333")
        self.log_box.pack(fill="both", expand=True, pady=20)
        self.log_box.configure(state="disabled")

    def setup_history(self):
        ctk.CTkLabel(self.hist_frame, text="DETECTION LOGS", font=(THEME["font_code"][0], 20, "bold")).pack(anchor="w", pady=(0, 20))
        cols = ("Time", "Company", "Source", "Details")
        self.tree = ttk.Treeview(self.hist_frame, columns=cols, show="headings", height=20)
        for col in cols: self.tree.heading(col, text=col.upper())
        self.tree.pack(fill="both", expand=True)
        CyberButton(self.hist_frame, text="REFRESH", command=self.load_history_data).pack(pady=10)

    def setup_settings(self):
        self.entries = {}
        fields = ["EMAIL_USER", "EMAIL_PASS", "TARGET_ID", "CHECK_INTERVAL"]
        for i, field in enumerate(fields):
            ctk.CTkLabel(self.set_frame, text=field.replace("_", " "), font=(THEME["font_code"][0], 12)).grid(row=i*2, column=0, sticky="w", pady=(10, 0))
            entry = CyberEntry(self.set_frame, width=400)
            if "PASS" in field: entry.configure(show="•")
            entry.insert(0, os.getenv(field, ""))
            entry.grid(row=i*2+1, column=0, sticky="w", pady=5)
            self.entries[field] = entry

        ctk.CTkButton(self.set_frame, text="SAVE CONFIG", fg_color=THEME["accent_green"], 
                      text_color="black", command=self.save_settings).grid(row=10, column=0, sticky="w", pady=30)

    def log(self, message):
        def _update():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _update)

    def toggle_monitoring(self):
        if not self.worker_running:
            self.worker_running = True
            self.btn_toggle.configure(text="STOP MONITOR", fg_color=THEME["error"])
            self.status_dot.configure(text_color=THEME["success"])
            self.status_text.configure(text="ONLINE", text_color=THEME["success"])
            threading.Thread(target=self.bg_loop, daemon=True).start()
        else:
            self.worker_running = False
            self.btn_toggle.configure(text="START MONITOR", fg_color=THEME["accent_blue"])
            self.status_dot.configure(text_color="grey")
            self.status_text.configure(text="OFFLINE", text_color="grey")

    def bg_loop(self):
        worker = MailWorker(self.log, self.trigger_alert)
        while self.worker_running:
            worker.run_check()
            interval = int(os.getenv("CHECK_INTERVAL", 30))
            for _ in range(interval):
                if not self.worker_running: break
                time.sleep(1)

    def trigger_alert(self, company):
        notification.notify(title="ID MATCH FOUND", message=f"Target detected in: {company}", timeout=10)

    def show_dashboard(self): self.switch_frame(self.dash_frame)
    def show_history(self): self.load_history_data(); self.switch_frame(self.hist_frame)
    def show_settings(self): self.switch_frame(self.set_frame)
    
    def switch_frame(self, frame):
        for f in [self.dash_frame, self.hist_frame, self.set_frame]: f.pack_forget()
        frame.pack(fill="both", expand=True)

    def save_settings(self):
        for key, entry in self.entries.items(): set_key(ENV_FILE, key, entry.get())
        load_dotenv(ENV_FILE, override=True)
        messagebox.showinfo("Success", "Configuration Saved.")

    def load_history_data(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for row in Database().get_all(): self.tree.insert("", "end", values=row)

if __name__ == "__main__":
    app = App()
    app.mainloop()